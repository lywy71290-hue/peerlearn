"""
Admin Blueprint — NITI Learn
Accessible only to users with is_admin=True.
"""
from functools import wraps
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort
from flask_login import login_required, current_user
from app import db
from models.user import User
from models.video import Video
from models.comment import Comment
from models.notification import Notification

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


# ── Decorator: require admin ──────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


def _notify_admins(message, link=""):
    """Send an in-app notification to all admins."""
    admins = User.query.filter_by(is_admin=True).all()
    for admin in admins:
        notif = Notification(user_id=admin.id, message=message, link=link)
        db.session.add(notif)
    db.session.commit()


# ── Dashboard ─────────────────────────────────────────────────────────────────
@admin_bp.route("/")
@login_required
@admin_required
def dashboard():
    total_users    = User.query.count()
    total_videos   = Video.query.filter_by(is_approved=True).count()
    pending_count  = Video.query.filter_by(is_approved=False).count()
    total_comments = Comment.query.count()
    recent_users   = User.query.order_by(User.created_at.desc()).limit(10).all()
    recent_videos  = Video.query.filter_by(is_approved=True).order_by(Video.created_at.desc()).limit(10).all()
    pending_videos = Video.query.filter_by(is_approved=False).order_by(Video.created_at.desc()).limit(5).all()
    return render_template(
        "admin/dashboard.html",
        total_users=total_users,
        total_videos=total_videos,
        pending_count=pending_count,
        total_comments=total_comments,
        recent_users=recent_users,
        recent_videos=recent_videos,
        pending_videos=pending_videos,
    )


# ── Pending Videos (Review Queue) ─────────────────────────────────────────────
@admin_bp.route("/pending")
@login_required
@admin_required
def pending():
    page = request.args.get("page", 1, type=int)
    pagination = (
        Video.query.filter_by(is_approved=False)
        .order_by(Video.created_at.desc())
        .paginate(page=page, per_page=15)
    )
    # Mark all admin notifications as read when visiting review page
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({"is_read": True})
    db.session.commit()
    return render_template("admin/pending.html", pagination=pagination)


# ── Approve Video ─────────────────────────────────────────────────────────────
@admin_bp.route("/videos/<int:video_id>/approve", methods=["POST"])
@login_required
@admin_required
def approve_video(video_id):
    video = Video.query.get_or_404(video_id)
    video.is_approved = True
    # Notify the uploader
    notif = Notification(
        user_id=video.user_id,
        message=f"✅ تمت الموافقة على فيديوك «{video.title}» ونُشر للجميع. حصلت على 10 نقاط! 🎉",
        link=f"/videos/{video.id}",
    )
    db.session.add(notif)

    # ── منح النقاط للرافع ──────────────────────────────────────────────────
    try:
        from models.rewards import PointTransaction
        uploader = video.author
        uploader.total_points = (uploader.total_points or 0) + 10
        tx = PointTransaction(
            user_id=uploader.id,
            points=10,
            reason=f"فيديو مقبول: {video.title}",
            video_id=video.id,
        )
        db.session.add(tx)
    except Exception as e:
        import logging; logging.getLogger(__name__).warning(f"Points grant failed: {e}")

    db.session.commit()
    # Send email to uploader
    try:
        from utils.email_notif import notify_uploader_approved
        notify_uploader_approved(
            uploader_email=video.author.email,
            uploader_name=video.author.username,
            video_title=video.title,
            video_id=video.id,
        )
    except Exception as e:
        import logging; logging.getLogger(__name__).warning(f"Approve email failed: {e}")
    flash(f"تمت الموافقة على «{video.title}» ونُشر.", "success")
    return redirect(request.referrer or url_for("admin.pending"))


# ── Reject Video ──────────────────────────────────────────────────────────────
@admin_bp.route("/videos/<int:video_id>/reject", methods=["POST"])
@login_required
@admin_required
def reject_video(video_id):
    video = Video.query.get_or_404(video_id)
    reason = request.form.get("reason", "").strip()
    # Notify the uploader
    msg = f"❌ تم رفض فيديوك «{video.title}»."
    if reason:
        msg += f" السبب: {reason}"
    notif = Notification(user_id=video.user_id, message=msg, link="")
    db.session.add(notif)
    # Send email to uploader
    try:
        from utils.email_notif import notify_uploader_rejected
        notify_uploader_rejected(
            uploader_email=video.author.email,
            uploader_name=video.author.username,
            video_title=video.title,
            reason=reason,
        )
    except Exception as e:
        import logging; logging.getLogger(__name__).warning(f"Reject email failed: {e}")
    # Delete the video record (and optionally from cloud)
    try:
        from utils.gcs import delete_blob_from_gcs
        delete_blob_from_gcs(video.gcs_url)
    except Exception:
        pass
    db.session.delete(video)
    db.session.commit()
    flash(f"تم رفض وحذف الفيديو «{video.title}».", "warning")
    return redirect(request.referrer or url_for("admin.pending"))


# ── Users List ────────────────────────────────────────────────────────────────
@admin_bp.route("/users")
@login_required
@admin_required
def users():
    page = request.args.get("page", 1, type=int)
    search = request.args.get("q", "").strip()
    query = User.query
    if search:
        query = query.filter(
            (User.username.ilike(f"%{search}%")) |
            (User.email.ilike(f"%{search}%"))
        )
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=20)
    return render_template("admin/users.html", pagination=pagination, search=search)


# ── Toggle Admin ──────────────────────────────────────────────────────────────
@admin_bp.route("/users/<int:user_id>/toggle-admin", methods=["POST"])
@login_required
@admin_required
def toggle_admin(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("لا يمكنك تعديل صلاحياتك الخاصة.", "warning")
        return redirect(url_for("admin.users"))
    user.is_admin = not user.is_admin
    db.session.commit()
    status = "مشرفاً" if user.is_admin else "متدرباً عادياً"
    flash(f"تم تغيير دور {user.username} إلى {status}.", "success")
    return redirect(url_for("admin.users"))


# ── Delete User ───────────────────────────────────────────────────────────────
@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("لا يمكنك حذف حسابك الخاص من هنا.", "warning")
        return redirect(url_for("admin.users"))
    db.session.delete(user)
    db.session.commit()
    flash(f"تم حذف المستخدم {user.username} بنجاح.", "success")
    return redirect(url_for("admin.users"))


# ── Videos List ───────────────────────────────────────────────────────────────
@admin_bp.route("/videos")
@login_required
@admin_required
def videos():
    page = request.args.get("page", 1, type=int)
    search = request.args.get("q", "").strip()
    query = Video.query
    if search:
        query = query.filter(Video.title.ilike(f"%{search}%"))
    pagination = query.order_by(Video.created_at.desc()).paginate(page=page, per_page=20)
    return render_template("admin/videos.html", pagination=pagination, search=search)


# ── Delete Video (admin) ──────────────────────────────────────────────────────
@admin_bp.route("/videos/<int:video_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_video(video_id):
    video = Video.query.get_or_404(video_id)
    title = video.title
    try:
        from utils.gcs import delete_blob_from_gcs
        delete_blob_from_gcs(video.gcs_url)
    except Exception:
        pass
    db.session.delete(video)
    db.session.commit()
    flash(f"تم حذف الفيديو «{title}» بنجاح.", "success")
    return redirect(url_for("admin.videos"))


# ── Import Trainees (CSV/Excel) ───────────────────────────────────────────────
@admin_bp.route("/import-trainees", methods=["GET", "POST"])
@login_required
@admin_required
def import_trainees():
    results = None
    if request.method == "POST":
        import io, csv
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Please upload a CSV or Excel file.", "danger")
            return render_template("admin/import_trainees.html", results=None)

        filename = file.filename.lower()
        added = 0
        skipped = 0
        errors = []

        try:
            if filename.endswith(".csv"):
                stream = io.StringIO(file.stream.read().decode("utf-8-sig"))
                reader = csv.DictReader(stream)
                rows = list(reader)
            elif filename.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(file.stream)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                flash("Unsupported file format. Please upload CSV or XLSX.", "danger")
                return render_template("admin/import_trainees.html", results=None)

            for i, row in enumerate(rows, start=2):
                # Normalize keys
                row = {k.strip().lower(): (str(v).strip() if v is not None else "") for k, v in row.items()}

                name        = row.get("name") or row.get("username") or row.get("full name") or row.get("trainee name") or ""
                national_id = row.get("national_id") or row.get("national id") or row.get("id") or row.get("رقم الهوية") or ""
                email       = row.get("email") or row.get("البريد") or ""
                program     = row.get("program") or row.get("البرنامج") or "academic"
                level       = row.get("level") or row.get("المستوى") or "Beginner"

                if not name or not national_id or not email:
                    errors.append(f"Row {i}: missing name, national_id, or email — skipped.")
                    skipped += 1
                    continue

                if not national_id.isdigit() or len(national_id) != 10:
                    errors.append(f"Row {i}: invalid national_id '{national_id}' — skipped.")
                    skipped += 1
                    continue

                if User.query.filter_by(national_id=national_id).first():
                    skipped += 1
                    continue

                if User.query.filter_by(email=email.lower()).first():
                    skipped += 1
                    continue

                user = User(
                    username=name,
                    email=email.lower(),
                    national_id=national_id,
                    program=program,
                    level=level,
                )
                user.set_password("Niti" + national_id)
                db.session.add(user)
                added += 1

            db.session.commit()
            results = {"added": added, "skipped": skipped, "errors": errors}
            flash(f"Import complete: {added} trainees added, {skipped} skipped.", "success")

        except Exception as e:
            flash(f"Error processing file: {e}", "danger")

    return render_template("admin/import_trainees.html", results=results)
